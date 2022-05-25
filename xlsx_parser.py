import enum
import re
import os

from db import DB
from log import Logger

from openpyxl import load_workbook
from sys import platform


class XLSXParser:

    def __init__(self, file):
        if platform == 'win32':
            folder = f'{os.getcwd()}\\schedule\\'
        else:
            folder = f'{os.getcwd()}/schedule/'
        self.__file = f'{folder}{file}'

    class __Days(enum.Enum):
        MONDAY = 1
        TUESDAY = 2
        WEDNESDAY = 3
        THURSDAY = 4
        FRIDAY = 5
        SATURDAY = 6

    class __Week(enum.Enum):
        EVEN = 0
        ODD = 1

    __file = ''
    __groups = list()
    __pattern = re.compile("([А-Я][А-Я][А-Я][А-Я]-[0-9][0-9]-[0-9][0-9])")

    def parse_file(self):

        Logger.info(self.__file)

        db = DB()

        wb = load_workbook(self.__file)
        ws = wb.active

        for row in ws.iter_rows(min_row=1, min_col=1, max_row=ws.max_row+1, max_col=ws.max_column+1):

            for cell in row:

                group_cell_value = str(cell.value)

                if self.__pattern.match(group_cell_value):

                    ind = min(len(group_cell_value), 10)
                    group = str(group_cell_value[:ind])

                    try:
                        db.add_group(group)
                    except StopIteration:
                        continue

                    column_index = cell.col_idx - 1

                    start = cell.row + 1
                    end = 75

                    para = 1
                    week = self.__Week.ODD.value

                    days_iterator = iter(self.__Days)
                    day = self.__Days.MONDAY

                    for i in range(start, end):

                        current_row = ws[i]

                        subject = current_row[column_index].value
                        subject_type = current_row[column_index + 1].value
                        teacher = current_row[column_index + 2].value
                        auditorium = current_row[column_index + 3].value

                        if subject is None:
                            subject = '-'
                        if subject_type is None:
                            subject_type = '-'
                        if teacher is None:
                            teacher = '-'
                        if auditorium is None:
                            auditorium = '-'

                        actual = i - 3

                        if actual % 12 == 0:  # Деление по дням
                            day = next(days_iterator)
                            para = 1

                        db.add_schedule(group, para, day.value, week, subject, subject_type, teacher, auditorium)

                        if actual % 2 != 0:  # Деление по неделям
                            para = para + 1
                            week = self.__Week.ODD.value
                        else:
                            week = self.__Week.EVEN.value

        db.close_connection()
        wb.close()
